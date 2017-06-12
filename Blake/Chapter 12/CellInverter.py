#! python3
# cellInverter.py - Inverts the row and columns. By this it means that it does a pivot,
# what good programmer doesn't know what a pivot is?

import openpyxl, sys

wb = openpyxl.load_workbook('InvertCells.xlsx')
sheet = wb.get_active_sheet()

sheetdata = []

for rowNum in range(1, sheet.get_highest_row() + 1):
    columnList = []
    for columnNum in range(1, sheet.get_highest_column() + 1):
        columnList = columnList + sheet.cell(row=rowNum, column=columnNum).value
    sheetdata[rowNum] = columnList

newWb = openpyxl.Workbook()

for columnNum in range(1, len(sheetdata)):
    for rowNum in range(1, len(sheetdata[columnNum])):
        sheet.cell(row=rowNum, column=columnNum).value = sheetdata[rowNum][columnNum]

newWb.save('InvertCells.xlsx')

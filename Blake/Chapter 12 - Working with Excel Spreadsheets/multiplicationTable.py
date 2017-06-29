#! python3
# multiplicationTable.py - Creates a multiplication table by taking in a single command line
# argument that tells it how many rows and columns. The program runs out of memory somewhere
# between 1000 and 2000

import openpyxl, sys
#from openpyxl.styles import Font, Style

if len(sys.argv) > 1:
    n = int(sys.argv[1])
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')

    # Set the labels and make them bold
    #fontObj = Font(bold=True)
    #styleObj = Style(font=fontObj)
    for num in range(1, n+1):
        sheet.cell(row=1, column=num+1).value = num
        #sheet.cell(row=1, column=num+1).style = styleObj
        sheet.cell(row=num+1, column=1).value = num
        #sheet.cell(row=num+1, column=1).style = styleObj

    for rowNum in range(1, n+1):
        for columnNum in range(1, n+1):
            sheet.cell(row=rowNum+1, column=columnNum+1).value = rowNum * columnNum

    wb.save('multiplicationTable.xlsx')

else:
    print("Must enter a number as a parameter")
